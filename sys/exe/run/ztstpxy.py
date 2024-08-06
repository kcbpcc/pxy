from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from openpyxl import load_workbook

class ExcelReaderApp(App):
    def build(self):
        layout = BoxLayout(orientation='vertical')
        scroll_view = ScrollView()
        content = BoxLayout(orientation='vertical', size_hint_y=None)
        content.bind(minimum_height=content.setter('height'))

        # Load Excel file
        workbook = load_workbook(filename='your_file.xlsx')
        sheet = workbook.active

        # Display Excel data
        for row in sheet.iter_rows(values_only=True):
            row_text = ' | '.join([str(cell) for cell in row])
            content.add_widget(Label(text=row_text, size_hint_y=None, height=40))

        scroll_view.add_widget(content)
        layout.add_widget(scroll_view)
        return layout

if __name__ == '__main__':
    ExcelReaderApp().run()
